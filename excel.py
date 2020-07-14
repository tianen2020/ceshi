from  openpyxl import load_workbook

#
def get_data():
    # 打开一个excel
    wb = load_workbook("pu.xlsx")
    sheet = wb.worksheets[0]
    # 操作单元格 行 列
    all_li= []
    for i in range(2, sheet.max_row+1):
        linli = []
        for j in range(1,sheet.max_column+1):
            # 获取value
                value = sheet.cell(i,j).value
                linli.append(value)

        all_li.append(linli)
    wb.close()
    return all_li

